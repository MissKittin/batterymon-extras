#!/usr/bin/env python3

# BatteryMon add-on that reads data from the archive
# and loads it into a spreadsheet
#
# Usage:
#  logs2xlsx.py /path/to/arch-journal --out=./output.xlsx
#  logs2xlsx.py /path/to/arch-journal > ./output.xlsx
#  logs2xlsx.py /path/to/arch-journal ./template.xlsx > ./output.xlsx
#  logs2xlsx.py /path/to/arch-journal ./template.xlsx --from=2025-04-03 > ./output.xlsx
#  logs2xlsx.py /path/to/arch-journal ./template.xlsx --from=2025-04-03 --to=2025-05-04 > ./output.xlsx
#  logs2xlsx.py --sheet-name "bt:01:23:45:67:89:AB"
#
# Additionally, you can add the argument --no-convert - this blocks conversion to a number (e.g. Balance="0010")
# Note: cell counting starts from 0
#  eg: --no-convert=3
#  eg: --no-convert=0,3,5
#
# Required packages: openpyxl

import sys
import gzip
import os
import re
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook

class argv_parser:
    def __init__(self, argv):
        self.__argv=[]
        self.__params={}

        for arg in argv:
            if arg[:2] == "--":
                arg_split=arg.split('=')
                self.__params[arg_split[0][2:]]=arg_split[1]

                continue

            self.__argv.append(arg)

    def __getitem__(self, index):
        return self.__argv[index]

    def __len__(self):
        return len(self.__argv)

    def arg(self, param):
        return self.__params.get(param)

def remove_array(array):
    return [v.strip("[]") for v in array]

def convert_nums(array, blacklist):
    result=[]

    for index, element in enumerate(array):
        if str(index) in blacklist:
            result.append(element)
            continue

        try:
            value=int(element)
        except ValueError:
            try:
                value=float(element)
            except ValueError:
                value=element

        result.append(value)

    return result

def sanitize_sheet_name(name):
    name=re.sub(r'[:\\/*?\[\]]', '_', name)

    if name.startswith("'"):
        name=name[1:]

    return name[:31]

if len(sys.argv) > 2 and sys.argv[1] == "--sheet-name":
    print(sanitize_sheet_name(sys.argv[2]))
    sys.exit(0)

argv=argv_parser(sys.argv)
template_path=argv[2] if len(argv) > 2 else None
from_ts=None
to_ts=None
no_convert=argv.arg("no-convert")

if no_convert is None:
    no_convert=[]
else:
    no_convert=no_convert.split(",")

if len(argv) < 2:
    print(argv[0]+" path/to/arch-journal [path/to/sheet.xlsx]")
    sys.exit(1)

if template_path:
    if not os.path.exists(template_path):
        print(template_path+" does not exists")
        sys.exit(1)

    wb=load_workbook(argv[2])
else:
    wb=Workbook()

if argv.arg("from"):
    try:
        from_ts=int(datetime.strptime(argv.arg("from"), "%Y-%m-%d").timestamp())
    except ValueError:
        print("Invalid date format in --from")
        sys.exit(1)

    if argv.arg("to"):
        try:
            to_ts=int(datetime.strptime(argv.arg("to"), "%Y-%m-%d").timestamp())
        except ValueError:
            print("Invalid date format in --to")
            sys.exit(1)

if not os.path.isdir(argv[1]):
    print(argv[1]+" is not a directory")
    sys.exit(1)

with os.scandir(argv[1]) as entries:
    for entry in entries:
        if not entry.is_file():
            continue

        entry_date=entry.name.split("_")[0]

        if from_ts:
            try:
                if int(datetime.strptime(entry_date, "%Y-%m-%d").timestamp()) < from_ts:
                    continue
            except ValueError:
                continue

        if to_ts:
            try:
                if int(datetime.strptime(entry_date, "%Y-%m-%d").timestamp()) > to_ts:
                    break
            except ValueError:
                continue

        with gzip.open(entry.path, "rt") as f:
            for line in f:
                values=line.strip().split()

                if values[2] != "OK":
                    continue

                mac=sanitize_sheet_name(values[3])
                del values[2:4] # OK MAC

                if mac not in wb.sheetnames:
                    wb.create_sheet(title=mac)

                wb[mac].append(convert_nums(remove_array(values), no_convert))

if argv.arg("out"):
    if os.path.exists(argv.arg("out")):
        print(argv.arg("out")+" already exist")
        sys.exit(1)

    wb.save(argv.arg("out"))
    sys.exit(0)

buf=BytesIO()
wb.save(buf)
buf.seek(0)
sys.stdout.buffer.write(buf.read())
